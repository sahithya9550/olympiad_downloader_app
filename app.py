import io
import os
import re
import zipfile
import mimetypes
from concurrent.futures import ThreadPoolExecutor, as_completed
from urllib.parse import urlparse, unquote

import pandas as pd
import requests
import streamlit as st
from openpyxl import load_workbook
from PyPDF2 import PdfMerger, PdfReader

APP_TITLE = "Universal Smart Bulk Downloader"
URL_PATTERN = re.compile(r"https?://[^\s\]\)\}\<\"']+", re.IGNORECASE)
GENERIC_HEADERS = {
    "name", "title", "topic", "subject", "category", "type", "file type",
    "document type", "resource type", "exam", "source", "year", "grade",
    "class", "level", "board", "course", "chapter", "unit", "notes",
    "status", "url", "link", "download", "file", "document"
}


def clean_text(value):
    if value is None or pd.isna(value):
        return ""
    return str(value).strip()


def clean_url(url):
    return str(url).strip().rstrip(".,);]}'\"")


def extract_urls_from_text(text):
    if not text:
        return []
    return list(dict.fromkeys(clean_url(u) for u in URL_PATTERN.findall(str(text))))


def normalize_column_name(value):
    return re.sub(r"\s+", " ", clean_text(value)).strip()


def make_safe_filename(value):
    value = unquote(clean_text(value))
    value = re.sub(r'[\\/*?:"<>|]', "_", value)
    value = re.sub(r"\s+", "_", value).strip("._ ")
    return (value or "file")[:150]


def looks_like_header(row_values):
    values = [normalize_column_name(v).lower() for v in row_values if clean_text(v)]
    if not values:
        return False
    score = 0
    for value in values:
        if value in GENERIC_HEADERS:
            score += 2
        if any(token in value for token in ("url", "link", "download", "file", "document")):
            score += 2
        if any(token in value for token in ("subject", "category", "topic", "title", "name", "year", "type")):
            score += 1
    return score >= 3


def find_header_row(raw_df):
    for row_index in range(min(len(raw_df), 30)):
        if looks_like_header(raw_df.iloc[row_index].tolist()):
            return row_index
    return None


def infer_link_type(column_name, url, nearby_text=""):
    text = f"{column_name} {url} {nearby_text}".lower()
    rules = [
        ("Solution", ["solution", "answer key", "answers", "key"]),
        ("Question Paper", ["question paper", "question", "exam paper", "test paper"]),
        ("Notes", ["notes", "study material", "material", "handout"]),
        ("Syllabus", ["syllabus", "curriculum"]),
        ("Video", ["video", "youtube", "youtu.be", "vimeo"]),
        ("Image", ["image", "photo", "picture", ".jpg", ".jpeg", ".png", ".webp"]),
        ("Spreadsheet", ["spreadsheet", "excel", ".xlsx", ".xls", ".csv"]),
        ("Presentation", ["presentation", "slides", ".ppt", ".pptx"]),
        ("Document", ["document", ".doc", ".docx", ".txt", ".rtf"]),
        ("PDF", ["pdf", ".pdf"]),
        ("Audio", ["audio", ".mp3", ".wav", ".m4a"]),
        ("Archive", ["archive", ".zip", ".rar", ".7z"]),
    ]
    for label, keywords in rules:
        if any(keyword in text for keyword in keywords):
            return label
    label = normalize_column_name(column_name)
    if label and not label.lower().startswith("unnamed"):
        return label
    return "File"


def choose_value(row_map, candidates, default=""):
    for candidate in candidates:
        for key, value in row_map.items():
            key_lower = key.lower()
            if candidate == key_lower or candidate in key_lower:
                cleaned = clean_text(value)
                if cleaned:
                    return cleaned
    return default


def row_metadata(row_map, sheet_name):
    subject = choose_value(row_map, ["subject", "category", "topic", "course", "department"], "")
    title = choose_value(row_map, ["title", "name", "exam", "resource", "document", "source", "book"], "")
    year = choose_value(row_map, ["year", "date", "session", "term"], "Unspecified")
    level = choose_value(row_map, ["grade", "class", "level", "board", "standard"], "")
    notes = choose_value(row_map, ["notes", "status", "description", "remarks"], "")

    # Headerless workbook fallback. For rows such as:
    # XII | Economics | Introductory Microeconomics | https://...
    # use the second value as Subject and the third as Book/Title.
    ordered_values = [clean_text(v) for v in row_map.values()]
    non_url_values = [v for v in ordered_values if v and not extract_urls_from_text(v)]
    if not subject:
        subject = non_url_values[1] if len(non_url_values) >= 2 else (non_url_values[0] if non_url_values else sheet_name)
    if not title:
        title = non_url_values[2] if len(non_url_values) >= 3 else (non_url_values[-1] if non_url_values else "Untitled")
    if not level and non_url_values:
        level = non_url_values[0]

    return subject or sheet_name, title or "Untitled", year, level, notes


def extract_hyperlinks(excel_bytes):
    """Return hyperlinks stored behind cell text, which pandas alone may miss."""
    workbook = load_workbook(io.BytesIO(excel_bytes), data_only=False, read_only=False)
    links = {}
    for worksheet in workbook.worksheets:
        sheet_links = {}
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.hyperlink and cell.hyperlink.target:
                    sheet_links[(cell.row, cell.column)] = clean_url(cell.hyperlink.target)
        links[worksheet.title] = sheet_links
    return links


def read_sheet(excel_bytes, sheet_name):
    raw_df = pd.read_excel(io.BytesIO(excel_bytes), sheet_name=sheet_name, header=None, dtype=str)
    header_row = find_header_row(raw_df)
    if header_row is None:
        df = raw_df.copy()
        df.columns = [f"Column {i + 1}" for i in range(len(df.columns))]
        excel_row_offset = 1
    else:
        df = pd.read_excel(io.BytesIO(excel_bytes), sheet_name=sheet_name, header=header_row, dtype=str)
        df.columns = [normalize_column_name(c) or f"Column {i + 1}" for i, c in enumerate(df.columns)]
        excel_row_offset = header_row + 2
    df = df.dropna(how="all").reset_index(drop=True)
    return df, header_row, excel_row_offset


def build_download_table(excel_bytes, source_workbook="Workbook"):
    excel = pd.ExcelFile(io.BytesIO(excel_bytes))
    hyperlinks = extract_hyperlinks(excel_bytes)
    all_rows, debug_rows = [], []

    for sheet_name in excel.sheet_names:
        df, header_row, excel_row_offset = read_sheet(excel_bytes, sheet_name)
        detected = 0

        for row_idx, row in df.iterrows():
            row_map = {str(col): row[col] for col in df.columns}
            subject, title, year, level, notes = row_metadata(row_map, sheet_name)
            row_text = " | ".join(clean_text(v) for v in row.tolist())
            seen = set()

            for col_idx, col_name in enumerate(df.columns, start=1):
                value = clean_text(row[col_name])
                urls = extract_urls_from_text(value)

                excel_row = excel_row_offset + row_idx
                hidden_link = hyperlinks.get(sheet_name, {}).get((excel_row, col_idx))
                if hidden_link:
                    urls.append(hidden_link)

                for url in dict.fromkeys(urls):
                    if url in seen:
                        continue
                    seen.add(url)
                    all_rows.append({
                        "Subject / Category": subject,
                        "Title / Source": title,
                        "Year / Date": year,
                        "Level / Group": level,
                        "File Type": infer_link_type(col_name, url, row_text),
                        "Source Column": col_name,
                        "URL": url,
                        "Notes": notes,
                        "Sheet": sheet_name,
                        "Source Workbook": source_workbook,
                    })
                    detected += 1

            # Final fallback: scan the complete row in case a URL spans unusual columns.
            for url in extract_urls_from_text(row_text):
                if url in seen:
                    continue
                all_rows.append({
                    "Subject / Category": subject,
                    "Title / Source": title,
                    "Year / Date": year,
                    "Level / Group": level,
                    "File Type": infer_link_type("", url, row_text),
                    "Source Column": "Row scan",
                    "URL": url,
                    "Notes": notes,
                    "Sheet": sheet_name,
                    "Source Workbook": source_workbook,
                })
                detected += 1

        debug_rows.append({
            "Source Workbook": source_workbook,
            "Sheet": sheet_name,
            "Header Row": header_row + 1 if header_row is not None else "No fixed header needed",
            "Rows Read": len(df),
            "Links Detected": detected,
            "Columns": ", ".join(map(str, df.columns)),
        })

    result_df = pd.DataFrame(all_rows)
    if not result_df.empty:
        result_df = result_df.drop_duplicates(subset=["URL", "Subject / Category", "Title / Source", "File Type", "Source Workbook"])
        result_df = result_df.sort_values(["Subject / Category", "Title / Source", "Year / Date", "File Type"]).reset_index(drop=True)
    return result_df, pd.DataFrame(debug_rows)


def get_extension(url, content_type, disposition=""):
    match = re.search(r'filename\*?=(?:UTF-8\'\')?["\']?([^"\';]+)', disposition, re.I)
    if match:
        ext = os.path.splitext(unquote(match.group(1)))[1]
        if ext:
            return ext
    ext = os.path.splitext(os.path.basename(urlparse(url).path))[1]
    if ext and len(ext) <= 10:
        return ext
    guessed = mimetypes.guess_extension((content_type or "").split(";")[0].strip())
    return guessed or ".bin"


def build_filename(item, content_type="", disposition=""):
    title = make_safe_filename(item.get("Title / Source", "file"))
    year = make_safe_filename(item.get("Year / Date", ""))
    file_type = make_safe_filename(item.get("File Type", "File"))
    ext = get_extension(item.get("URL", ""), content_type, disposition)
    parts = [p for p in (title, year, file_type) if p and p.lower() not in ("unspecified", "untitled")]
    return make_safe_filename("_".join(parts) or "file") + ext


def download_file(item):
    try:
        response = requests.get(
            item["URL"],
            headers={"User-Agent": "Mozilla/5.0 UniversalBulkDownloader/2.0", "Accept": "*/*"},
            timeout=60,
            allow_redirects=True,
        )
        response.raise_for_status()
        content_type = response.headers.get("Content-Type", "")
        disposition = response.headers.get("Content-Disposition", "")
        return {**item, "Status": "Success", "Filename": build_filename(item, content_type, disposition), "Content Type": content_type, "Error": "", "Content": response.content}
    except requests.exceptions.RequestException as exc:
        return {**item, "Status": "Failed", "Filename": "", "Error": str(exc), "Content": None}


def add_file_to_zip(zip_file, used_paths, zip_path, content):
    base, ext = os.path.splitext(zip_path)
    final_path, counter = zip_path, 1
    while final_path in used_paths:
        final_path = f"{base}_{counter}{ext}"
        counter += 1
    used_paths.add(final_path)
    zip_file.writestr(final_path, content)
    return final_path


def create_report_excel(results):
    report_df = pd.DataFrame([{k: v for k, v in item.items() if k != "Content"} for item in results])
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        report_df.to_excel(writer, index=False, sheet_name="Download Report")
    return output.getvalue()


def merge_pdfs(pdf_items):
    merger, added = PdfMerger(), 0
    for item in pdf_items:
        try:
            content = item.get("Content")
            PdfReader(io.BytesIO(content))
            merger.append(io.BytesIO(content))
            added += 1
        except Exception:
            continue
    if not added:
        merger.close()
        return None
    output = io.BytesIO()
    merger.write(output)
    merger.close()
    return output.getvalue()


def extract_text_from_pdf(content):
    try:
        reader = PdfReader(io.BytesIO(content))
        return "".join(f"\n\n--- Page {i} ---\n{page.extract_text() or ''}" for i, page in enumerate(reader.pages, 1)).strip()
    except Exception as exc:
        return f"Text extraction failed: {exc}"


def is_zip_download(item):
    filename = clean_text(item.get("Filename", "")).lower()
    content_type = clean_text(item.get("Content Type", "")).lower()
    return filename.endswith(".zip") or "application/zip" in content_type or "x-zip" in content_type


def archive_stem(item):
    url_name = os.path.basename(urlparse(item.get("URL", "")).path)
    stem = os.path.splitext(unquote(url_name))[0]
    return make_safe_filename(stem or item.get("Title / Source") or "Book")


def book_folder_name(item):
    title = clean_text(item.get("Title / Source", ""))
    if title and title.lower() not in {"untitled", "file"}:
        return make_safe_filename(title)
    return archive_stem(item)


def add_download_to_master_zip(zf, used_paths, item):
    """Add a direct file or flatten a downloaded ZIP into Subject/Book/."""
    subject = make_safe_filename(item.get("Subject / Category") or "General")
    book = book_folder_name(item)
    base_folder = f"{subject}/{book}"
    content = item.get("Content")

    if is_zip_download(item):
        try:
            with zipfile.ZipFile(io.BytesIO(content), "r") as source_zip:
                extracted = 0
                for member in source_zip.infolist():
                    if member.is_dir():
                        continue
                    # Ignore archive-internal folders and place every actual file
                    # directly under Subject/Book/.
                    member_name = make_safe_filename(os.path.basename(member.filename))
                    if not member_name:
                        continue
                    member_content = source_zip.read(member)
                    final_path = add_file_to_zip(
                        zf, used_paths, f"{base_folder}/{member_name}", member_content
                    )
                    extracted += 1
                if extracted:
                    item["Zip Path"] = f"{base_folder}/ ({extracted} extracted files)"
                    return
        except (zipfile.BadZipFile, RuntimeError, OSError) as exc:
            item["Archive Warning"] = f"Could not extract ZIP; saved original archive: {exc}"

    item["Zip Path"] = add_file_to_zip(
        zf, used_paths, f"{base_folder}/{item['Filename']}", content
    )


def create_zip(results, include_originals=True, include_merged=False, include_text=False):
    """Create one master ZIP arranged as Subject / Book / actual files."""
    output, used_paths = io.BytesIO(), set()
    successful = [r for r in results if r.get("Status") == "Success" and r.get("Content")]

    with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED, allowZip64=True) as zf:
        # Keep the batch report at the ZIP root.
        zf.writestr("download_report.xlsx", create_report_excel(results))

        if include_originals:
            for item in successful:
                add_download_to_master_zip(zf, used_paths, item)

        if include_text:
            for item in successful:
                if item.get("Filename", "").lower().endswith(".pdf"):
                    subject = make_safe_filename(item.get("Subject / Category") or "General")
                    base_name = os.path.splitext(item["Filename"])[0]
                    path = f"{subject}/{base_name}_text.txt"
                    add_file_to_zip(
                        zf,
                        used_paths,
                        path,
                        extract_text_from_pdf(item["Content"]).encode("utf-8", "ignore"),
                    )

        if include_merged:
            subjects = sorted({item.get("Subject / Category") or "General" for item in successful})
            for subject in subjects:
                items = [
                    item for item in successful
                    if (item.get("Subject / Category") or "General") == subject
                    and item.get("Filename", "").lower().endswith(".pdf")
                ]
                merged = merge_pdfs(items)
                if merged:
                    safe_subject = make_safe_filename(subject)
                    add_file_to_zip(
                        zf,
                        used_paths,
                        f"{safe_subject}/{safe_subject}_merged.pdf",
                        merged,
                    )

    output.seek(0)
    return output.getvalue()


def main():
    st.set_page_config(page_title=APP_TITLE, page_icon="📦", layout="wide")
    st.title("📦 Universal Smart Bulk Downloader")
    st.write(
        "Upload one or more Excel workbooks containing links. The app reads every sheet and row, "
        "automatically identifies subjects, categories, titles, years, file types, and hidden Excel hyperlinks, "
        "then creates one master ZIP arranged as Subject / Book Title / extracted files. Downloaded ZIP packages are automatically unpacked, and their internal folders are flattened."
    )

    with st.sidebar:
        st.header("Master ZIP Options")
        include_originals = st.checkbox("Original downloaded files", True)
        include_merged = st.checkbox("Merge direct PDFs by subject", False)
        include_text = st.checkbox("Extract text from PDFs", False)
        show_debug = st.checkbox("Show detection details", False)
        max_workers = st.slider("Parallel downloads", min_value=2, max_value=16, value=8, help="Higher values can be faster, but some websites may throttle or block too many simultaneous requests.")

    uploaded_files = st.file_uploader(
        "Upload one or more Excel workbooks",
        type=["xlsx", "xlsm", "xls"],
        accept_multiple_files=True,
    )
    if not uploaded_files:
        st.info("Upload one or more Excel files to begin.")
        return

    download_tables = []
    debug_tables = []
    workbook_errors = []

    for uploaded_file in uploaded_files:
        try:
            file_df, file_debug_df = build_download_table(
                uploaded_file.getvalue(),
                source_workbook=uploaded_file.name,
            )
            if not file_df.empty:
                download_tables.append(file_df)
            if not file_debug_df.empty:
                debug_tables.append(file_debug_df)
        except Exception as exc:
            workbook_errors.append({"Workbook": uploaded_file.name, "Error": str(exc)})

    download_df = (
        pd.concat(download_tables, ignore_index=True)
        if download_tables else pd.DataFrame()
    )
    debug_df = (
        pd.concat(debug_tables, ignore_index=True)
        if debug_tables else pd.DataFrame()
    )

    if workbook_errors:
        st.warning(f"{len(workbook_errors)} workbook(s) could not be read.")
        st.dataframe(pd.DataFrame(workbook_errors), use_container_width=True, hide_index=True)

    if show_debug:
        st.subheader("Detection Details")
        st.dataframe(debug_df, use_container_width=True, hide_index=True)

    if download_df.empty:
        st.error("No downloadable http:// or https:// links were found in the uploaded workbooks.")
        return

    st.success(
        f"Detected {len(download_df)} downloadable links across "
        f"{download_df['Source Workbook'].nunique()} workbook(s) and "
        f"{download_df['Sheet'].nunique()} sheet name(s)."
    )
    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Workbooks", download_df["Source Workbook"].nunique())
    c2.metric("Links", len(download_df))
    c3.metric("Categories", download_df["Subject / Category"].nunique())
    c4.metric("Titles / Sources", download_df["Title / Source"].nunique())
    c5.metric("File Types", download_df["File Type"].nunique())

    workbooks = sorted(download_df["Source Workbook"].astype(str).unique())
    categories = sorted(download_df["Subject / Category"].astype(str).unique())
    types = sorted(download_df["File Type"].astype(str).unique())
    sheets = sorted(download_df["Sheet"].astype(str).unique())
    a, b, c, d = st.columns(4)
    with a:
        selected_workbooks = st.multiselect("Source Workbooks", workbooks, default=workbooks)
    with b:
        selected_categories = st.multiselect("Subjects / Categories", categories, default=categories)
    with c:
        selected_types = st.multiselect("File Types", types, default=types)
    with d:
        selected_sheets = st.multiselect("Sheets", sheets, default=sheets)

    filtered = download_df[
        download_df["Source Workbook"].isin(selected_workbooks)
        & download_df["Subject / Category"].isin(selected_categories)
        & download_df["File Type"].isin(selected_types)
        & download_df["Sheet"].isin(selected_sheets)
    ].reset_index(drop=True)

    st.subheader("Detected Files")
    st.dataframe(filtered, use_container_width=True, hide_index=True)
    if filtered.empty:
        st.warning("No links match the selected filters.")
        return
    if not any((include_originals, include_merged, include_text)):
        st.warning("Select at least one output option.")
        return

    if st.button("⬇️ Build One Master ZIP", type="primary"):
        records = filtered.to_dict("records")
        results = [None] * len(records)
        progress, status = st.progress(0), st.empty()

        # Download several independent URLs at the same time. Threading is
        # effective here because network waiting, not Python computation, is
        # the main bottleneck. Results are written back in the original order.
        completed = 0
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            future_to_index = {
                executor.submit(download_file, item): index
                for index, item in enumerate(records)
            }
            for future in as_completed(future_to_index):
                index = future_to_index[future]
                try:
                    results[index] = future.result()
                except Exception as exc:
                    item = records[index]
                    results[index] = {
                        **item,
                        "Status": "Failed",
                        "Filename": "",
                        "Error": str(exc),
                        "Content": None,
                    }
                completed += 1
                status.write(f"Downloaded {completed}/{len(records)} files")
                progress.progress(completed / len(records))

        zip_bytes = create_zip(results, include_originals, include_merged, include_text)
        success = sum(r.get("Status") == "Success" for r in results)
        st.success(f"Completed: {success} successful, {len(results) - success} failed.")
        report = pd.DataFrame([{k: v for k, v in r.items() if k != "Content"} for r in results])
        st.dataframe(report, use_container_width=True, hide_index=True)
        st.download_button("⬇️ Download One Master ZIP", zip_bytes, "all_downloads_subject_book_extracted.zip", "application/zip")


if __name__ == "__main__":
    main()